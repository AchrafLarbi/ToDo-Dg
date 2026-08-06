"""Blueprint pour la gestion des projets."""
from datetime import datetime, date
from flask import Blueprint, render_template, request, redirect, url_for, flash
from app.models import database

projects_bp = Blueprint('projects', __name__)

@projects_bp.route('/projets')
def projets():
    projets_list = database.list_projets()
    return render_template('projects/projets.html', projets=projets_list, today=date.today().isoformat())

@projects_bp.route('/projets/ajouter', methods=['POST'])
def ajouter_projet():
    name = request.form['name'].strip()
    description = request.form.get('description', '').strip()
    start_date = request.form.get('start_date', '').strip()
    deadline = request.form.get('deadline', '').strip()
    status = request.form.get('status', 'En cours')
    if name:
        database.create_projet(name, description, start_date, deadline, status)
        flash('Projet créé.', 'success')
    return redirect(url_for('projects.projets'))

@projects_bp.route('/projets/<int:id>')
def detail_projet(id):
    projet = database.get_projet(id)
    if not projet:
        flash('Projet introuvable.', 'danger')
        return redirect(url_for('projects.projets'))
    tasks = database.list_taches_by_projet(id)
    return render_template('projects/projet_detail.html', projet=projet, tasks=tasks, today=date.today().isoformat())

@projects_bp.route('/projets/<int:id>/modifier', methods=['POST'])
def modifier_projet(id):
    name = request.form['name'].strip()
    description = request.form.get('description', '').strip()
    start_date = request.form.get('start_date', '').strip()
    deadline = request.form.get('deadline', '').strip()
    status = request.form.get('status', 'En cours')
    if name:
        database.update_projet(id, name, description, start_date, deadline, status)
        flash('Projet mis à jour.', 'success')
    return redirect(url_for('projects.detail_projet', id=id))

@projects_bp.route('/projets/<int:id>/supprimer', methods=['POST'])
def supprimer_projet(id):
    database.delete_projet(id)
    flash('Projet et tâches associées supprimés.', 'success')
    return redirect(url_for('projects.projets'))


@projects_bp.route('/projets/vue')
def vue_par_projet():
    today = date.today().isoformat()
    projets_list = database.list_projets()
    analytics = database.get_projects_analytics(today)

    selected_project_id = request.args.get('project_id', type=int)
    selected_project = None
    project_tasks = []
    project_stats = None

    if selected_project_id:
        selected_project = database.get_projet(selected_project_id)

    if not selected_project and projets_list:
        selected_project = projets_list[0]
        selected_project_id = selected_project['id']

    if selected_project:
        project_tasks = database.list_taches_by_projet(selected_project_id)
        # Calculate stats for selected project
        total = len(project_tasks)
        actives = sum(1 for t in project_tasks if t.get('status') not in ('Terminee', 'Cloturee', 'Terminé'))
        closed = sum(1 for t in project_tasks if t.get('status') in ('Terminee', 'Cloturee', 'Terminé'))
        overdue = sum(1 for t in project_tasks if t.get('due_date') and t.get('due_date') < today and t.get('status') not in ('Terminee', 'Cloturee', 'Terminé'))
        completion_pct = round((closed / total * 100)) if total > 0 else 0

        project_stats = {
            'total': total,
            'actives': actives,
            'closed': closed,
            'overdue': overdue,
            'completion_pct': completion_pct,
        }

    return render_template(
        'projects/vue_par_projet.html',
        projets=projets_list,
        selected_project=selected_project,
        project_tasks=project_tasks,
        project_stats=project_stats,
        analytics=analytics,
        today=today,
    )


@projects_bp.route('/projets/importer-excel', methods=['POST'])
def importer_excel():
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Veuillez sélectionner un fichier Excel (.xlsx, .xls) ou CSV.', 'danger')
        return redirect(url_for('projects.vue_par_projet'))

    filename = file.filename.lower()
    rows_data = []

    try:
        if filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
            headers = [str(cell.value or '').strip().lower() for cell in next(sheet.iter_rows(max_row=1))]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {}
                    for idx, val in enumerate(row):
                        if idx < len(headers):
                            row_dict[headers[idx]] = str(val).strip() if val is not None else ''
                    rows_data.append(row_dict)

        elif filename.endswith('.csv'):
            import csv, io
            stream = io.StringIO(file.stream.read().decode('utf-8-sig', errors='ignore'), newline=None)
            reader = csv.DictReader(stream)
            for row in reader:
                clean_row = {str(k or '').strip().lower(): str(v or '').strip() for k, v in row.items()}
                if any(clean_row.values()):
                    rows_data.append(clean_row)
        else:
            flash('Format de fichier non supporté. Utilisez .xlsx, .xls ou .csv', 'danger')
            return redirect(url_for('projects.vue_par_projet'))

    except Exception as exc:
        flash(f"Erreur lors de la lecture du fichier : {exc}", 'danger')
        return redirect(url_for('projects.vue_par_projet'))

    if not rows_data:
        flash('Le fichier envoyé ne contient aucune donnée valide.', 'warning')
        return redirect(url_for('projects.vue_par_projet'))

    # Process rows
    imported_count = 0
    created_projects = set()

    for r in rows_data:
        # Helper to get value matching key aliases
        def get_val(aliases):
            for a in aliases:
                for k in r.keys():
                    if a in k:
                        return r[k]
            return ''

        project_name = get_val(['projet', 'project'])
        title = get_val(['tache', 'tâches', 'titre', 'title', 'subject', 'libelle'])
        responsable = get_val(['responsable', 'collaborateur', 'assign', 'email', 'nom'])
        echeance_raw = get_val(['echeance', 'échéance', 'date', 'due'])

        if not title:
            continue

        # Find or create project
        project_id = database.find_or_create_project_by_name(project_name) if project_name else None
        if project_name and project_id:
            created_projects.add(project_name.strip())

        # Find or create collaborator
        collab_id = database.find_or_create_collaborator_by_name_or_email(responsable) if responsable else None

        # Parse date
        due_date = None
        if echeance_raw:
            raw_str = str(echeance_raw).split(' ')[0] # Trim time part if datetime string
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    dt = datetime.strptime(raw_str, fmt)
                    due_date = dt.date().isoformat()
                    break
                except ValueError:
                    pass

        tid = database.create_tache(
            title=title,
            description=f"Importé depuis Excel (Projet : {project_name or 'Aucun'})",
            project_id=project_id,
            collaborator_id=collab_id,
            priority='Normale',
            sensitivity='Normale',
            due_date=due_date,
        )

        if collab_id:
            from app.services import send_task_notification_async
            send_task_notification_async(tid)

        imported_count += 1

    flash(f"Importation terminée avec succès ! {imported_count} tâche(s) créées sur {len(created_projects)} projet(s).", 'success')
    return redirect(url_for('projects.vue_par_projet'))


@projects_bp.route('/projets/modele-excel')
def modele_excel():
    import io, openpyxl
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import_Taches"

    headers = ['projet', 'taches', 'responsable', 'echeances']
    ws.append(headers)

    sample_row = ['Projet Exemple', 'Tâche Exemple - Titre de la tâche', 'Collaborateur Exemple', '2026-08-30']
    ws.append(sample_row)

    # Format column width & headers
    for col_idx, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="modele_importation_taches.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


